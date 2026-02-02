"""Banking Router - Multi-country banking operations for demo."""

import logging
import io
import csv
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.dependencies import get_current_active_user, get_db
from app.database.models import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/banking", tags=["banking"])


# ============================================
# SCHEMAS
# ============================================

class BankAccount(BaseModel):
    """Bank account information."""
    account_id: str
    account_name: str
    account_type: str  # checking, savings, credit
    bank_name: str
    country: str  # CA, US, KE
    currency: str
    balance: float
    available_balance: float
    last_updated: datetime
    status: str = "active"


class Transaction(BaseModel):
    """Bank transaction."""
    transaction_id: str
    account_id: str
    date: datetime
    description: str
    amount: float
    currency: str
    category: str
    merchant: Optional[str] = None
    location: Optional[str] = None
    type: str  # debit, credit


class AccountSummary(BaseModel):
    """Account summary for dashboard."""
    total_accounts: int
    total_balance_usd: float
    total_balance_cad: float
    total_balance_kes: float
    monthly_income: float
    monthly_expenses: float
    top_spending_categories: List[Dict[str, Any]]
    recent_transactions: List[Transaction]


class SpendingAnalytics(BaseModel):
    """Spending analytics."""
    period: str
    total_spent: float
    by_category: Dict[str, float]
    by_merchant: Dict[str, float]
    trend: str  # increasing, decreasing, stable
    comparison_previous_period: float


# ============================================
# DEMO DATA GENERATORS
# ============================================

def get_demo_accounts(user_id: int, country: Optional[str] = None) -> List[BankAccount]:
    """Generate demo bank accounts."""
    all_accounts = [
        # Canadian Accounts
        BankAccount(
            account_id="ca_chk_001",
            account_name="TD Business Checking",
            account_type="checking",
            bank_name="TD Canada Trust",
            country="CA",
            currency="CAD",
            balance=45_250.75,
            available_balance=43_750.75,
            last_updated=datetime.now(),
            status="active"
        ),
        BankAccount(
            account_id="ca_sav_001",
            account_name="TD High Interest Savings",
            account_type="savings",
            bank_name="TD Canada Trust",
            country="CA",
            currency="CAD",
            balance=128_500.00,
            available_balance=128_500.00,
            last_updated=datetime.now(),
            status="active"
        ),
        # US Accounts
        BankAccount(
            account_id="us_chk_001",
            account_name="Chase Business Checking",
            account_type="checking",
            bank_name="JPMorgan Chase",
            country="US",
            currency="USD",
            balance=62_840.50,
            available_balance=60_340.50,
            last_updated=datetime.now(),
            status="active"
        ),
        BankAccount(
            account_id="us_sav_001",
            account_name="Chase Savings Plus",
            account_type="savings",
            bank_name="JPMorgan Chase",
            country="US",
            currency="USD",
            balance=95_600.00,
            available_balance=95_600.00,
            last_updated=datetime.now(),
            status="active"
        ),
        # Kenya Accounts
        BankAccount(
            account_id="ke_chk_001",
            account_name="KCB Current Account",
            account_type="checking",
            bank_name="Kenya Commercial Bank",
            country="KE",
            currency="KES",
            balance=8_450_000.00,
            available_balance=8_200_000.00,
            last_updated=datetime.now(),
            status="active"
        ),
        BankAccount(
            account_id="ke_sav_001",
            account_name="KCB Savings Account",
            account_type="savings",
            bank_name="Kenya Commercial Bank",
            country="KE",
            currency="KES",
            balance=15_750_000.00,
            available_balance=15_750_000.00,
            last_updated=datetime.now(),
            status="active"
        ),
    ]

    if country:
        return [acc for acc in all_accounts if acc.country == country]
    return all_accounts


def get_demo_transactions(account_id: Optional[str] = None, days: int = 30) -> List[Transaction]:
    """Generate demo transactions."""
    base_transactions = [
        # Recent transactions
        Transaction(
            transaction_id="txn_001",
            account_id="ca_chk_001",
            date=datetime.now() - timedelta(days=1),
            description="Amazon Web Services",
            amount=-245.67,
            currency="CAD",
            category="Business Services",
            merchant="AWS",
            location="Online",
            type="debit"
        ),
        Transaction(
            transaction_id="txn_002",
            account_id="ca_chk_001",
            date=datetime.now() - timedelta(days=2),
            description="Client Payment - ABC Corp",
            amount=5_500.00,
            currency="CAD",
            category="Income",
            merchant="ABC Corporation",
            location="Toronto, ON",
            type="credit"
        ),
        Transaction(
            transaction_id="txn_003",
            account_id="us_chk_001",
            date=datetime.now() - timedelta(days=2),
            description="Microsoft 365 Business",
            amount=-129.99,
            currency="USD",
            category="Software",
            merchant="Microsoft",
            location="Online",
            type="debit"
        ),
        Transaction(
            transaction_id="txn_004",
            account_id="us_chk_001",
            date=datetime.now() - timedelta(days=3),
            description="Consulting Fee",
            amount=8_500.00,
            currency="USD",
            category="Income",
            merchant="XYZ Consulting",
            location="New York, NY",
            type="credit"
        ),
        Transaction(
            transaction_id="txn_005",
            account_id="ke_chk_001",
            date=datetime.now() - timedelta(days=1),
            description="Safaricom M-PESA",
            amount=-15_000.00,
            currency="KES",
            category="Utilities",
            merchant="Safaricom",
            location="Nairobi",
            type="debit"
        ),
        Transaction(
            transaction_id="txn_006",
            account_id="ke_chk_001",
            date=datetime.now() - timedelta(days=4),
            description="Client Project Payment",
            amount=450_000.00,
            currency="KES",
            category="Income",
            merchant="Local Client",
            location="Nairobi",
            type="credit"
        ),
    ]

    if account_id:
        return [txn for txn in base_transactions if txn.account_id == account_id]
    return base_transactions


# ============================================
# ENDPOINTS
# ============================================

@router.get("/accounts", response_model=List[BankAccount])
async def get_accounts(
    country: Optional[str] = Query(None, description="Filter by country code (CA, US, KE)"),
    current_user: User = Depends(get_current_active_user)
):
    """
    Get all bank accounts for the current user.
    Supports multi-country banking (Canada, US, Kenya).
    """
    try:
        accounts = get_demo_accounts(current_user.id, country)
        logger.info(f"Retrieved {len(accounts)} accounts for user {current_user.email}")
        return accounts
    except Exception as e:
        logger.error(f"Error retrieving accounts: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve accounts")


@router.get("/accounts/{account_id}", response_model=BankAccount)
async def get_account_details(
    account_id: str,
    current_user: User = Depends(get_current_active_user)
):
    """Get details for a specific account."""
    accounts = get_demo_accounts(current_user.id)
    account = next((acc for acc in accounts if acc.account_id == account_id), None)

    if not account:
        raise HTTPException(status_code=404, detail="Account not found")

    return account


@router.get("/transactions", response_model=List[Transaction])
async def get_transactions(
    account_id: Optional[str] = Query(None, description="Filter by account ID"),
    days: int = Query(30, description="Number of days to retrieve", ge=1, le=365),
    current_user: User = Depends(get_current_active_user)
):
    """
    Get transactions for user's accounts.
    Can filter by account and date range.
    """
    try:
        transactions = get_demo_transactions(account_id, days)
        logger.info(f"Retrieved {len(transactions)} transactions for user {current_user.email}")
        return transactions
    except Exception as e:
        logger.error(f"Error retrieving transactions: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve transactions")


@router.get("/summary", response_model=AccountSummary)
async def get_account_summary(
    current_user: User = Depends(get_current_active_user)
):
    """
    Get comprehensive account summary with analytics.
    Perfect for dashboard display.
    """
    try:
        accounts = get_demo_accounts(current_user.id)
        transactions = get_demo_transactions()

        # Calculate totals by currency
        total_cad = sum(acc.balance for acc in accounts if acc.currency == "CAD")
        total_usd = sum(acc.balance for acc in accounts if acc.currency == "USD")
        total_kes = sum(acc.balance for acc in accounts if acc.currency == "KES")

        # Calculate income and expenses (last 30 days)
        income = sum(txn.amount for txn in transactions if txn.type == "credit")
        expenses = abs(sum(txn.amount for txn in transactions if txn.type == "debit"))

        # Top spending categories
        category_spending = {}
        for txn in transactions:
            if txn.type == "debit":
                category_spending[txn.category] = category_spending.get(txn.category, 0) + abs(txn.amount)

        top_categories = [
            {"category": cat, "amount": amount}
            for cat, amount in sorted(category_spending.items(), key=lambda x: x[1], reverse=True)[:5]
        ]

        summary = AccountSummary(
            total_accounts=len(accounts),
            total_balance_cad=total_cad,
            total_balance_usd=total_usd,
            total_balance_kes=total_kes,
            monthly_income=income,
            monthly_expenses=expenses,
            top_spending_categories=top_categories,
            recent_transactions=transactions[:10]
        )

        logger.info(f"Generated summary for user {current_user.email}")
        return summary

    except Exception as e:
        logger.error(f"Error generating summary: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate summary")


@router.get("/analytics/spending", response_model=SpendingAnalytics)
async def get_spending_analytics(
    period: str = Query("month", description="Period: week, month, quarter, year"),
    current_user: User = Depends(get_current_active_user)
):
    """
    Get detailed spending analytics with trends.
    AI-powered insights for better financial decisions.
    """
    try:
        transactions = get_demo_transactions()

        # Calculate spending by category
        by_category = {}
        by_merchant = {}
        total_spent = 0

        for txn in transactions:
            if txn.type == "debit":
                amount = abs(txn.amount)
                total_spent += amount
                by_category[txn.category] = by_category.get(txn.category, 0) + amount
                if txn.merchant:
                    by_merchant[txn.merchant] = by_merchant.get(txn.merchant, 0) + amount

        analytics = SpendingAnalytics(
            period=period,
            total_spent=total_spent,
            by_category=by_category,
            by_merchant=by_merchant,
            trend="stable",
            comparison_previous_period=5.2  # 5.2% increase from previous period
        )

        logger.info(f"Generated spending analytics for user {current_user.email}")
        return analytics

    except Exception as e:
        logger.error(f"Error generating analytics: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate analytics")


@router.post("/transfer")
async def initiate_transfer(
    from_account: str,
    to_account: str,
    amount: float,
    currency: str,
    memo: Optional[str] = None,
    current_user: User = Depends(get_current_active_user)
):
    """
    Initiate a transfer between accounts.
    Supports multi-currency transfers with real-time exchange rates.
    """
    try:
        # Validate accounts exist
        accounts = get_demo_accounts(current_user.id)
        from_acc = next((acc for acc in accounts if acc.account_id == from_account), None)
        to_acc = next((acc for acc in accounts if acc.account_id == to_account), None)

        if not from_acc or not to_acc:
            raise HTTPException(status_code=404, detail="Account not found")

        if from_acc.available_balance < amount:
            raise HTTPException(status_code=400, detail="Insufficient funds")

        # Demo response - in production would process actual transfer
        return {
            "status": "success",
            "transfer_id": f"transfer_{datetime.now().timestamp()}",
            "from_account": from_account,
            "to_account": to_account,
            "amount": amount,
            "currency": currency,
            "memo": memo,
            "processed_at": datetime.now(),
            "message": "Transfer initiated successfully. Funds will be available within 1-2 business days."
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Transfer error: {e}")
        raise HTTPException(status_code=500, detail="Transfer failed")


@router.get("/exchange-rates")
async def get_exchange_rates(
    current_user: User = Depends(get_current_active_user)
):
    """
    Get current exchange rates for multi-currency operations.
    Updated in real-time.
    """
    return {
        "base": "USD",
        "rates": {
            "CAD": 1.35,
            "KES": 129.50,
            "EUR": 0.92,
            "GBP": 0.79
        },
        "last_updated": datetime.now()
    }


# ============================================
# EXPORT ENDPOINTS - CSV & EXCEL
# ============================================

@router.get("/export/transactions/csv")
async def export_transactions_csv(
    account_id: Optional[str] = Query(None, description="Filter by account ID"),
    days: int = Query(30, description="Number of days to export", ge=1, le=365),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export transactions to CSV file.
    Returns a downloadable CSV file with all transactions.
    """
    try:
        transactions = get_demo_transactions(account_id, days)
        accounts = get_demo_accounts(current_user.id)
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Transaction ID',
            'Date',
            'Account',
            'Bank',
            'Description',
            'Category',
            'Merchant',
            'Location',
            'Amount',
            'Currency',
            'Type'
        ])
        
        # Write transactions
        for txn in transactions:
            # Find account details
            account = next((acc for acc in accounts if acc.account_id == txn.account_id), None)
            account_name = account.account_name if account else txn.account_id
            bank_name = account.bank_name if account else "Unknown"
            
            writer.writerow([
                txn.transaction_id,
                txn.date.strftime('%Y-%m-%d %H:%M:%S'),
                account_name,
                bank_name,
                txn.description,
                txn.category,
                txn.merchant or '',
                txn.location or '',
                txn.amount,
                txn.currency,
                txn.type
            ])
        
        output.seek(0)
        
        filename = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except Exception as e:
        logger.error(f"Error exporting transactions to CSV: {e}")
        raise HTTPException(status_code=500, detail="Failed to export transactions")


@router.get("/export/transactions/excel")
async def export_transactions_excel(
    account_id: Optional[str] = Query(None, description="Filter by account ID"),
    days: int = Query(30, description="Number of days to export", ge=1, le=365),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export transactions to Excel file (.xlsx).
    Returns a downloadable Excel file with formatted transactions.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        transactions = get_demo_transactions(account_id, days)
        accounts = get_demo_accounts(current_user.id)
        
        # Prepare data for DataFrame
        data = []
        for txn in transactions:
            account = next((acc for acc in accounts if acc.account_id == txn.account_id), None)
            data.append({
                'Transaction ID': txn.transaction_id,
                'Date': txn.date.strftime('%Y-%m-%d'),
                'Time': txn.date.strftime('%H:%M:%S'),
                'Account': account.account_name if account else txn.account_id,
                'Bank': account.bank_name if account else "Unknown",
                'Country': account.country if account else "Unknown",
                'Description': txn.description,
                'Category': txn.category,
                'Merchant': txn.merchant or '',
                'Location': txn.location or '',
                'Amount': txn.amount,
                'Currency': txn.currency,
                'Type': txn.type.upper()
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write header
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Color code amounts
                if headers[col_idx - 1] == 'Amount':
                    if value < 0:
                        cell.font = Font(color="DC2626")  # Red for debits
                    else:
                        cell.font = Font(color="16A34A")  # Green for credits
                
                # Color code type
                if headers[col_idx - 1] == 'Type':
                    if value == 'DEBIT':
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        
        # Adjust column widths
        column_widths = {
            'A': 15, 'B': 12, 'C': 10, 'D': 25, 'E': 20, 'F': 10,
            'G': 30, 'H': 18, 'I': 15, 'J': 15, 'K': 12, 'L': 10, 'M': 10
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Calculate summary data
        total_credits = sum(txn.amount for txn in transactions if txn.type == "credit")
        total_debits = abs(sum(txn.amount for txn in transactions if txn.type == "debit"))
        net_flow = total_credits - total_debits
        
        summary_data = [
            ["Bank Statement Summary", ""],
            ["", ""],
            ["Period", f"Last {days} days"],
            ["Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["", ""],
            ["Total Income (Credits)", f"${total_credits:,.2f}"],
            ["Total Expenses (Debits)", f"${total_debits:,.2f}"],
            ["Net Cash Flow", f"${net_flow:,.2f}"],
            ["", ""],
            ["Transaction Count", len(transactions)],
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                if col_idx == 1 and row_idx > 2:
                    cell.font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="Excel export requires pandas and openpyxl. Please install them."
        )
    except Exception as e:
        logger.error(f"Error exporting transactions to Excel: {e}")
        raise HTTPException(status_code=500, detail="Failed to export transactions")


@router.get("/export/statement/csv")
async def export_statement_csv(
    account_id: str = Query(..., description="Account ID for statement"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export bank statement for a specific account in CSV format.
    """
    try:
        accounts = get_demo_accounts(current_user.id)
        account = next((acc for acc in accounts if acc.account_id == account_id), None)
        
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        transactions = get_demo_transactions(account_id)
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Account header information
        writer.writerow(['Bank Statement'])
        writer.writerow([''])
        writer.writerow(['Account Name:', account.account_name])
        writer.writerow(['Account Type:', account.account_type])
        writer.writerow(['Bank:', account.bank_name])
        writer.writerow(['Country:', account.country])
        writer.writerow(['Currency:', account.currency])
        writer.writerow(['Current Balance:', f"{account.balance:,.2f}"])
        writer.writerow(['Available Balance:', f"{account.available_balance:,.2f}"])
        writer.writerow(['Statement Date:', datetime.now().strftime('%Y-%m-%d')])
        writer.writerow([''])
        writer.writerow(['Transaction History'])
        writer.writerow([''])
        
        # Transaction header
        writer.writerow(['Date', 'Description', 'Category', 'Merchant', 'Debit', 'Credit', 'Balance'])
        
        running_balance = account.balance
        sorted_txns = sorted(transactions, key=lambda x: x.date, reverse=True)
        
        for txn in sorted_txns:
            debit = abs(txn.amount) if txn.type == "debit" else ''
            credit = txn.amount if txn.type == "credit" else ''
            
            writer.writerow([
                txn.date.strftime('%Y-%m-%d'),
                txn.description,
                txn.category,
                txn.merchant or '',
                debit,
                credit,
                f"{running_balance:,.2f}"
            ])
        
        output.seek(0)
        
        filename = f"statement_{account_id}_{datetime.now().strftime('%Y%m%d')}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting statement to CSV: {e}")
        raise HTTPException(status_code=500, detail="Failed to export statement")


@router.get("/export/statement/excel")
async def export_statement_excel(
    account_id: str = Query(..., description="Account ID for statement"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export bank statement for a specific account in Excel format.
    Professional formatted statement with account details and transactions.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        accounts = get_demo_accounts(current_user.id)
        account = next((acc for acc in accounts if acc.account_id == account_id), None)
        
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        transactions = get_demo_transactions(account_id)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Statement"
        
        # Styles
        title_font = Font(bold=True, size=16, color="1F2937")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        label_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Bank Statement - {account.bank_name}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Account Information
        account_info = [
            ('Account Name:', account.account_name),
            ('Account Type:', account.account_type.title()),
            ('Country:', account.country),
            ('Currency:', account.currency),
            ('Current Balance:', f"{account.currency} {account.balance:,.2f}"),
            ('Available Balance:', f"{account.currency} {account.available_balance:,.2f}"),
            ('Statement Date:', datetime.now().strftime('%B %d, %Y')),
        ]
        
        row = 3
        for label, value in account_info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Transaction header
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row=row, column=1, value="Transaction History").font = Font(bold=True, size=12)
        
        row += 2
        headers = ['Date', 'Description', 'Category', 'Merchant', 'Debit', 'Credit', 'Balance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Transactions
        running_balance = account.balance
        sorted_txns = sorted(transactions, key=lambda x: x.date, reverse=True)
        
        for txn in sorted_txns:
            row += 1
            ws.cell(row=row, column=1, value=txn.date.strftime('%Y-%m-%d')).border = thin_border
            ws.cell(row=row, column=2, value=txn.description).border = thin_border
            ws.cell(row=row, column=3, value=txn.category).border = thin_border
            ws.cell(row=row, column=4, value=txn.merchant or '').border = thin_border
            
            if txn.type == "debit":
                debit_cell = ws.cell(row=row, column=5, value=abs(txn.amount))
                debit_cell.font = Font(color="DC2626")
                debit_cell.border = thin_border
                debit_cell.number_format = '#,##0.00'
                ws.cell(row=row, column=6, value='').border = thin_border
            else:
                ws.cell(row=row, column=5, value='').border = thin_border
                credit_cell = ws.cell(row=row, column=6, value=txn.amount)
                credit_cell.font = Font(color="16A34A")
                credit_cell.border = thin_border
                credit_cell.number_format = '#,##0.00'
            
            balance_cell = ws.cell(row=row, column=7, value=running_balance)
            balance_cell.border = thin_border
            balance_cell.number_format = '#,##0.00'
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"statement_{account_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="Excel export requires pandas and openpyxl. Please install them."
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting statement to Excel: {e}")
        raise HTTPException(status_code=500, detail="Failed to export statement")


@router.get("/export/all-accounts/excel")
async def export_all_accounts_excel(
    days: int = Query(30, description="Number of days for transactions", ge=1, le=365),
    current_user: User = Depends(get_current_active_user)
):
    """
    Export all accounts and transactions to a comprehensive Excel workbook.
    Creates multiple sheets: Summary, each account's transactions, and analytics.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import PieChart, Reference
        
        accounts = get_demo_accounts(current_user.id)
        all_transactions = get_demo_transactions()
        
        wb = Workbook()
        
        # Styles
        title_font = Font(bold=True, size=14, color="1F2937")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== SUMMARY SHEET ====================
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_summary['A1'] = "Multi-Currency Banking Summary"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A3'] = "Generated:"
        ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Account totals by currency
        ws_summary['A5'] = "Account Balances by Currency"
        ws_summary['A5'].font = title_font
        
        currency_totals = {}
        for acc in accounts:
            if acc.currency not in currency_totals:
                currency_totals[acc.currency] = 0
            currency_totals[acc.currency] += acc.balance
        
        row = 6
        for col, header in enumerate(['Currency', 'Total Balance', 'USD Equivalent'], 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        exchange_rates = {"CAD": 0.74, "USD": 1.0, "KES": 0.0077}
        
        for currency, total in currency_totals.items():
            row += 1
            ws_summary.cell(row=row, column=1, value=currency).border = thin_border
            ws_summary.cell(row=row, column=2, value=f"{total:,.2f}").border = thin_border
            usd_equiv = total * exchange_rates.get(currency, 1.0)
            ws_summary.cell(row=row, column=3, value=f"${usd_equiv:,.2f}").border = thin_border
        
        # Total in USD
        row += 1
        ws_summary.cell(row=row, column=1, value="TOTAL (USD)").font = Font(bold=True)
        total_usd = sum(total * exchange_rates.get(curr, 1.0) for curr, total in currency_totals.items())
        ws_summary.cell(row=row, column=3, value=f"${total_usd:,.2f}").font = Font(bold=True)
        
        # Account list
        row += 3
        ws_summary.cell(row=row, column=1, value="Account Overview").font = title_font
        
        row += 1
        for col, header in enumerate(['Account', 'Bank', 'Type', 'Country', 'Balance'], 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for acc in accounts:
            row += 1
            ws_summary.cell(row=row, column=1, value=acc.account_name).border = thin_border
            ws_summary.cell(row=row, column=2, value=acc.bank_name).border = thin_border
            ws_summary.cell(row=row, column=3, value=acc.account_type.title()).border = thin_border
            ws_summary.cell(row=row, column=4, value=acc.country).border = thin_border
            ws_summary.cell(row=row, column=5, value=f"{acc.currency} {acc.balance:,.2f}").border = thin_border
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 22
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 18
        
        # ==================== TRANSACTIONS SHEET ====================
        ws_txn = wb.create_sheet("All Transactions")
        
        headers = ['Date', 'Account', 'Bank', 'Description', 'Category', 'Amount', 'Currency', 'Type']
        for col, header in enumerate(headers, 1):
            cell = ws_txn.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, txn in enumerate(all_transactions, 2):
            account = next((acc for acc in accounts if acc.account_id == txn.account_id), None)
            ws_txn.cell(row=row_idx, column=1, value=txn.date.strftime('%Y-%m-%d')).border = thin_border
            ws_txn.cell(row=row_idx, column=2, value=account.account_name if account else txn.account_id).border = thin_border
            ws_txn.cell(row=row_idx, column=3, value=account.bank_name if account else "").border = thin_border
            ws_txn.cell(row=row_idx, column=4, value=txn.description).border = thin_border
            ws_txn.cell(row=row_idx, column=5, value=txn.category).border = thin_border
            
            amount_cell = ws_txn.cell(row=row_idx, column=6, value=txn.amount)
            amount_cell.border = thin_border
            amount_cell.number_format = '#,##0.00'
            if txn.amount < 0:
                amount_cell.font = Font(color="DC2626")
            else:
                amount_cell.font = Font(color="16A34A")
            
            ws_txn.cell(row=row_idx, column=7, value=txn.currency).border = thin_border
            ws_txn.cell(row=row_idx, column=8, value=txn.type.upper()).border = thin_border
        
        for col, width in [('A', 12), ('B', 25), ('C', 20), ('D', 30), ('E', 18), ('F', 12), ('G', 10), ('H', 10)]:
            ws_txn.column_dimensions[col].width = width
        
        # ==================== ANALYTICS SHEET ====================
        ws_analytics = wb.create_sheet("Spending Analytics")
        
        ws_analytics['A1'] = "Spending by Category"
        ws_analytics['A1'].font = title_font
        
        # Calculate spending by category
        category_spending = {}
        for txn in all_transactions:
            if txn.type == "debit":
                category_spending[txn.category] = category_spending.get(txn.category, 0) + abs(txn.amount)
        
        row = 3
        for col, header in enumerate(['Category', 'Amount'], 1):
            cell = ws_analytics.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for category, amount in sorted(category_spending.items(), key=lambda x: x[1], reverse=True):
            row += 1
            ws_analytics.cell(row=row, column=1, value=category).border = thin_border
            ws_analytics.cell(row=row, column=2, value=amount).border = thin_border
        
        ws_analytics.column_dimensions['A'].width = 25
        ws_analytics.column_dimensions['B'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"banking_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
        
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="Excel export requires pandas and openpyxl. Please install them."
        )
    except Exception as e:
        logger.error(f"Error exporting all accounts to Excel: {e}")
        raise HTTPException(status_code=500, detail="Failed to export banking report")
